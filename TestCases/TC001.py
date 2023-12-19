from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from keyboard import press_and_release
import keyboard
from selenium.webdriver import ActionChains
import random
from bs4 import BeautifulSoup
import requests
from urllib.request import urlopen
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter
import HtmlTestRunner
from datetime import date
import pyautogui
import csv



PATH = "/usr/local/bin/chromedriver"

DriverService = Service(PATH)

driver = webdriver.Chrome(service = DriverService)

driver.get("https://industriamaquiladora.com/index.php")

#Print URL of Chrome
str_URL = driver.current_url

print("URL: "+str_URL)

#Closing pop up window
ClosePopUp = driver.find_element(By.CLASS_NAME,"close")
ClosePopUp.click()

#Click on proveedores button
ProveedoresButton = driver.find_element(By.XPATH,'/html/body/section[3]/div/div/div[2]/div')
ProveedoresButton.click()
#time.sleep(2)

print (driver.current_url)

#How many iterations will be performed

ExcelFilePath = 'Excel file'
LoadWbIterations = load_workbook("/Users/alancastillo-osuna/Documents/Final Exam/Data/ProvidersData.xlsx")
ExcelSheetIterations = LoadWbIterations['Sheet2']
IterationRow = ExcelSheetIterations['A2'].value
print(IterationRow)


for x in range(IterationRow):
    print("Working on")

    #Read container for Proveedores and get total items, with class name col-md-6
    #not all items from the list have this class name
    list = driver.find_elements(By.CLASS_NAME,"col-md-6")
    size = len(list)
    print('Proveedores col-md-6: '+str(size))

    #driver.back()

    #randomly selects an item to enter
    SelectedProvider = random.randint(1,size)
    print (SelectedProvider)
    print (driver.current_url)

    #pick the item index from excel, all full xpath are there, 74 in total
    #To access the value for a specific cell:
    ExcelFilePath = 'Excel file'
    LoadWb = load_workbook("/Users/alancastillo-osuna/Documents/Final Exam/Data/ProvidersData.xlsx")
    ExcelSheet=LoadWb['Sheet1']
    Row = "C" + str(SelectedProvider)
    RowVal = str(Row)
    print(Row)
    print(RowVal)
    ExcelRow = ExcelSheet[RowVal].value
    ExcelRowVal = str(ExcelRow)
    print(ExcelRow)
    print(ExcelRowVal)


    #Click on random item from size, comes from locators, a list of all full xpaths
    #Provider = driver.find_element(By.XPATH,"/html/body/section[3]/div/div/div[1]/div[3]/div/div")
    Provider = driver.find_element(By.XPATH, ExcelRowVal).click()
    print (driver.current_url)


    #click to select item inside providers
    #fix xpath for  select              /////////////////
    #listProvider = driver.find_elements(By.CLASS_NAME,"col-md-2")
    #sizeP = len(listProvider)
    #print (sizeP)
    #print (driver.current_url)

    #Click on provider to select
    driver.find_element(By.XPATH, "/html/body/section[4]/div[2]/div/div[1]/div[1]").click()
    print (driver.current_url)

    #Save screenshot
    ScreenCapture = str(x) + "_Screen.png"
    myScreenshot = pyautogui.screenshot()
    myScreenshot.save(r'/Users/alancastillo-osuna/Documents/Final Exam/Evidence/'+ScreenCapture)

    #Pick a provider from the specific category and detects data from the tables

    url = driver.current_url
    page = requests.get(url)
    page.text
    #print(page.text)

    soup = BeautifulSoup(page.text, 'html.parser')
    page = requests.get(url)

    inner_div = soup.find_all('div', class_="col-md-8")
    print(inner_div)

    #Save data on Excel file

    pathEvidence = '/Users/alancastillo-osuna/Documents/Final Exam//Evidence/Final_results.csv'
    soup1 = BeautifulSoup(page.text, 'html.parser')

    with open(pathEvidence, "w", encoding="utf-8-sig", newline='') as csv_file:
        w = csv.writer(csv_file, delimiter = ",", quoting=csv.QUOTE_MINIMAL)
        w.writerow([i.text for i in soup.select('.form-group label')])
        w.writerow([i['value'] for i in soup.select('input.aspNetDisabled')])


    #time.sleep(5)
    driver.back()
    driver.back()
    print (driver.current_url)

else:
    print("Completed iterations")

#Evidence saved on HTML file
todayDate = date.today()
print("Today's date:", todayDate)
OutputFolder = "/Users/alancastillo-osuna/Documents/Final Exam/"
OutputFile = "_Evidence.html"
output1 = OutputFolder + str(todayDate)+ OutputFile
testRunner=HtmlTestRunner.HTMLTestRunner(output = output1)
print(output1)

#time.sleep(2)
driver.close()
driver.quit()
